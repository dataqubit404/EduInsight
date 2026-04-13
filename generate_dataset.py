import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import random

np.random.seed(42)
random.seed(42)

subjects = [
    "Essentials Of Data Science",
    "Software Engineering",
    "Deep Learning",
    "Technical Training"
]

first_names = [
    "Aarav","Aditya","Akash","Amit","Ananya","Anjali","Arjun","Aryan","Ayesha",
    "Deepak","Deepika","Divya","Farhan","Fatima","Gaurav","Harshit","Ishaan",
    "Isha","Jiya","Kabir","Kavya","Kiran","Kunal","Lakshmi","Manish","Meera",
    "Mohammad","Mohit","Nandini","Neha","Nikhil","Nisha","Om","Palak","Pankaj",
    "Pooja","Priya","Rahul","Ravi","Ritika","Rohan","Rohit","Sachin","Sakshi",
    "Sandeep","Sanjana","Sanjay","Sara","Shivam","Shreya","Shubham","Simran",
    "Sneha","Soham","Sonali","Suresh","Tanvi","Uday","Varun","Vikram","Yash","Zara"
]

def generate_student_data(subject, n=40):
    rows = []
    names_pool = random.sample(first_names, min(n, len(first_names)))
    if n > len(first_names):
        names_pool += [f"Student_{i}" for i in range(n - len(first_names))]

    for i in range(n):
        study_hours = round(np.random.uniform(1.0, 8.0), 1)
        attendance  = round(np.random.uniform(50, 100), 1)

        # Midterm 0-50
        midterm_raw = round(np.clip(
            20 + study_hours * 3 + np.random.normal(0, 5), 0, 50))
        midterm_scaled = round(midterm_raw * 20 / 50, 1)   # scale 0-20

        quiz1 = round(np.clip(study_hours * 2 + np.random.normal(0, 2), 0, 20), 1)
        quiz2 = round(np.clip(study_hours * 1.8 + np.random.normal(0, 3), 0, 20), 1)
        best_quiz = max(quiz1, quiz2)

        lab_viva = round(np.clip(study_hours * 0.8 + np.random.normal(0, 1), 0, 10), 1)
        project  = round(np.clip(study_hours * 0.9 + np.random.normal(0, 1.5), 0, 10), 1)

        # End-term 0-100 raw → scaled 0-40
        endterm_raw = round(np.clip(
            30 + study_hours * 4.5 + np.random.normal(0, 8), 0, 100))
        endterm_scaled = round(endterm_raw * 40 / 100, 1)

        total = round(midterm_scaled + best_quiz + lab_viva + project + endterm_scaled, 1)
        total = min(total, 100)

        if total >= 85:   grade = "A+"
        elif total >= 75: grade = "A"
        elif total >= 65: grade = "B+"
        elif total >= 55: grade = "B"
        elif total >= 45: grade = "C+"
        elif total >= 35: grade = "C"
        elif total >= 30: grade = "D"
        else:             grade = "F"

        rows.append({
            "Student Name": names_pool[i],
            "Subject": subject,
            "Study Hours (per day)": study_hours,
            "Attendance (%)": attendance,
            "Midterm Marks (0-50)": midterm_raw,
            "Midterm Scaled (0-20)": midterm_scaled,
            "Quiz 1 (0-20)": quiz1,
            "Quiz 2 (0-20)": quiz2,
            "Best Quiz (0-20)": best_quiz,
            "Lab + Viva (0-10)": lab_viva,
            "Project (0-10)": project,
            "End-Term Raw (0-100)": endterm_raw,
            "End-Term Scaled (0-40)": endterm_scaled,
            "Total Marks (0-100)": total,
            "Grade": grade
        })
    return pd.DataFrame(rows)

# ── Build workbook ──────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUB_FILL   = PatternFill("solid", fgColor="2E86AB")
SUB_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT_FILL   = PatternFill("solid", fgColor="EBF5FB")
NORM_FONT  = Font(name="Arial", size=10)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")
thin       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

grade_colors = {
    "A+": "27AE60", "A": "2ECC71", "B+": "F39C12",
    "B": "E67E22", "C+": "E74C3C", "C": "C0392B",
    "D": "8E44AD", "F": "2C3E50"
}

all_dfs = []

for subject in subjects:
    df = generate_student_data(subject, 40)
    all_dfs.append(df)
    sheet_name = subject[:25].replace(" ", "_")
    ws = wb.create_sheet(title=sheet_name)

    # Title row
    ws.merge_cells("A1:O1")
    ws["A1"] = f"📊 {subject} — Student Performance Dataset"
    ws["A1"].font  = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill  = PatternFill("solid", fgColor="0D2137")
    ws["A1"].alignment = CENTER

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CENTER; cell.border = BORDER

    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORM_FONT; cell.border = BORDER
            cell.alignment = CENTER if col_idx > 1 else LEFT
            cell.fill = fill
        # Color grade cell
        grade_val = row["Grade"]
        grade_cell = ws.cell(row=row_idx, column=15)
        grade_cell.fill  = PatternFill("solid", fgColor=grade_colors.get(grade_val, "CCCCCC"))
        grade_cell.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    col_widths = [22,28,20,16,20,20,14,14,16,16,14,20,20,18,10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 40
    for r in range(3, 44):
        ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A3"

# ── Summary sheet ────────────────────────────────────────────────────────────
all_df = pd.concat(all_dfs, ignore_index=True)
ws_sum = wb.create_sheet(title="Summary", index=0)
ws_sum.merge_cells("A1:F1")
ws_sum["A1"] = "📋 Student Performance Summary — All Subjects"
ws_sum["A1"].font  = Font(name="Arial", bold=True, color="FFFFFF", size=13)
ws_sum["A1"].fill  = PatternFill("solid", fgColor="0D2137")
ws_sum["A1"].alignment = CENTER

sum_headers = ["Subject","Total Students","Avg Study Hours",
               "Avg Total Marks","Pass Rate (%)","Top Grade (A+) Count"]
for c, h in enumerate(sum_headers, 1):
    cell = ws_sum.cell(row=2, column=c, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = CENTER; cell.border = BORDER

for r, subject in enumerate(subjects, 3):
    sub = all_df[all_df["Subject"] == subject]
    vals = [
        subject,
        len(sub),
        round(sub["Study Hours (per day)"].mean(), 2),
        round(sub["Total Marks (0-100)"].mean(), 2),
        round((sub["Grade"] != "F").sum() / len(sub) * 100, 1),
        int((sub["Grade"] == "A+").sum())
    ]
    for c, v in enumerate(vals, 1):
        cell = ws_sum.cell(row=r, column=c, value=v)
        cell.font = NORM_FONT; cell.border = BORDER
        cell.alignment = CENTER if c > 1 else LEFT
        cell.fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

for c, w in zip(range(1,7), [32,18,18,18,16,22]):
    ws_sum.column_dimensions[get_column_letter(c)].width = w
ws_sum.row_dimensions[1].height = 28; ws_sum.row_dimensions[2].height = 36

path = "/home/claude/student_dashboard/data/student_performance.xlsx"
wb.save(path)

# Also save raw CSV for ML
all_df.to_csv("/home/claude/student_dashboard/data/student_performance.csv", index=False)
print(f"✅ Dataset saved — {len(all_df)} students across {len(subjects)} subjects")
print(all_df["Grade"].value_counts().sort_index())
